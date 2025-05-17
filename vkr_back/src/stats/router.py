from fastapi.responses import StreamingResponse
from fastapi.routing import APIRouter
from fastapi import Depends, HTTPException
from datetime import date, datetime, timedelta
from src.stats.utils import hist_plot_by_difficulty
from src.database import get_db_session
from fastapi import Path, Query
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from src.stats.utils import (
    get_employee_tasks,
    get_team_tasks,
    get_project_tasks,
    calculate_average_status_durations,
    count_status_transitions,
    average_completion_time,
    average_completion_time_by_difficulty,
    percentage_completed_on_time,
    count_active_tasks,
    get_amount_of_done_tasks_by_user_id_and_period,
    get_teamlead_users,
    get_id_name_by_user_id,
    get_avg_difficulty_by_user_id_and_period
)

router = APIRouter(
    prefix='/stats',
    tags=['stats']
)

@router.get('/hist-plot-by-difficulty')
def get_hist_plot_by_difficulty(project_id: int | None = None, user_id: int = Query(...), period_start: datetime = Query(...), period_end: datetime = Query(...)):
    with get_db_session() as session:
        return hist_plot_by_difficulty(project_id, user_id, period_start, period_end, session)
    
    
@router.get("/employee", summary="Метрики по сотруднику")
async def employee_metrics(
    user_id: int,
    start_date: datetime,
    end_date: datetime
):
    """
    Возвращает аналитические метрики для сотрудника по заданному периоду:
      - Среднее время в статусах;
      - Количество переходов по статусам;
      - Среднее время выполнения задач;
      - Среднее время выполнения задач по сложности;
      - Процент задач, выполненных вовремя;
      - Количество активных задач.
    """
    with get_db_session() as session:
        tasks = get_employee_tasks(session, user_id, start_date, end_date)
        if not tasks:
            raise HTTPException(status_code=404, detail="Задачи для указанного сотрудника не найдены")
    
    metrics = {
        "average_status_durations": calculate_average_status_durations(tasks),
        "status_transitions": count_status_transitions(tasks, start_date, end_date),
        "average_completion_time": average_completion_time(tasks),
        "average_completion_time_by_difficulty": average_completion_time_by_difficulty(tasks),
        "percentage_completed_on_time": percentage_completed_on_time(tasks),
        "active_tasks_count": count_active_tasks(tasks)
    }
    print(metrics)
    
    return metrics


@router.get("/team", summary="Метрики по команде")
async def team_metrics(
    boss_id: int,
    start_date: datetime,
    end_date: datetime,

):
    """
    Возвращает аналитические метрики для команды (сотрудники с одинаковым boss_id)
    по заданному периоду:
      - Среднее время в статусах;
      - Количество переходов по статусам;
      - Среднее время выполнения задач;
      - Среднее время выполнения задач по сложности;
      - Процент задач, выполненных вовремя;
      - Количество активных задач.
    """
    with get_db_session() as session:
        tasks = get_team_tasks(session, boss_id, start_date, end_date)
        if not tasks:
            return {}
            raise HTTPException(status_code=404, detail="Задачи для указанной команды не найдены")
    
    metrics = {
        "average_status_durations": calculate_average_status_durations(tasks),
        "status_transitions": count_status_transitions(tasks, start_date, end_date),
        "average_completion_time": average_completion_time(tasks),
        "average_completion_time_by_difficulty": average_completion_time_by_difficulty(tasks),
        "percentage_completed_on_time": percentage_completed_on_time(tasks),
        "active_tasks_count": count_active_tasks(tasks)
    }
    return metrics


@router.get("/project", summary="Метрики по проекту")
async def project_metrics(
    project_id: int,
    start_date: datetime,
    end_date: datetime,
):
    """
    Возвращает аналитические метрики для проекта по заданному периоду:
      - Среднее время в статусах;
      - Количество переходов по статусам;
      - Среднее время выполнения задач;
      - Среднее время выполнения задач по сложности;
      - Процент задач, выполненных вовремя;
      - Количество активных задач.
    """
    with get_db_session() as session:   
        tasks = get_project_tasks(session, project_id, start_date, end_date)
        if not tasks:
            raise HTTPException(status_code=404, detail="Задачи для указанного проекта не найдены")
    print(tasks)
    metrics = {
        "average_status_durations": calculate_average_status_durations(tasks),
        "status_transitions": count_status_transitions(tasks, start_date, end_date),
        "average_completion_time": average_completion_time(tasks),
        "average_completion_time_by_difficulty": average_completion_time_by_difficulty(tasks),
        "percentage_completed_on_time": percentage_completed_on_time(tasks),
        "active_tasks_count": count_active_tasks(tasks)
    }
    return metrics



@router.get('/report')
async def report(
    start_date: datetime,
    end_date: datetime,
    user_id: int | None = None,
    project_id: int | None = None,
):
    # Проверка корректности дат
    if start_date > end_date:
        raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
    
    # Формирование списка дат (только даты, с интервалом 1 день)
    num_days = (end_date.date() - start_date.date()).days + 1
    date_list = [start_date.date() + timedelta(days=i) for i in range(num_days)]
    
    with get_db_session() as session:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет задач"
        
        # Записываем заголовки: первый столбец – "Сотрудник", затем по одному столбцу для каждой даты
        cell = ws.cell(row=1, column=1, value="")
        cell.font = Font(size=14)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for idx, single_date in enumerate(date_list, start=2):
            cell = ws.cell(row=2, column=idx, value=single_date.strftime("%Y-%m-%d"))
            cell.font = Font(size=14)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell = ws.cell(row=2, column=len(date_list)+2, value="Всего")   
        cell.font = Font(size=14)
        cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        if user_id and project_id:
            # Заполняем строки: для каждого сотрудника – подсчет выполненных задач по дням
            row_num = 3
            teamleads = get_teamlead_users(user_id=user_id, session=session)
            if teamleads:
                users = {user: get_teamlead_users(user_id=user[0], session=session) for user in teamleads}
                print(users)
                for teamlead, users in users.items():
                    cell = ws.cell(row=row_num, column=1, value=teamlead[1])
                    cell.font = Font(bold=True, size=14)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    total_amount = 0
                    total_difficulty = 0
                    
                    for col_idx, single_date in enumerate(date_list, start=2):
                        day_start = datetime.combine(single_date, datetime.min.time())
                        day_end = day_start + timedelta(days=1)
                        amount = get_amount_of_done_tasks_by_user_id_and_period(project_id=project_id, user_id=teamlead[0], period_start=day_start, period_end=day_end, session=session)
                        difficulty = get_avg_difficulty_by_user_id_and_period(project_id=project_id, user_id=teamlead[0], period_start=day_start, period_end=day_end, session=session)
                        cell = ws.cell(row=row_num, column=col_idx, value=f'{amount} / {difficulty}')
                        cell.font = Font(size=14)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        total_amount += amount
                        total_difficulty += difficulty
                    cell = ws.cell(row=row_num, column=len(date_list)+2, value=f'{total_amount} / {total_difficulty}')
                    cell.font = Font(size=14)
                    cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    row_num += 1
                    for user in users:
                        cell = ws.cell(row=row_num, column=1, value=user[1])
                        cell.font = Font(size=14)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        total_amount = 0
                        total_difficulty = 0
                        for col_idx, single_date in enumerate(date_list, start=2):
                            day_start = datetime.combine(single_date, datetime.min.time())
                            day_end = day_start + timedelta(days=1)
                            amount = get_amount_of_done_tasks_by_user_id_and_period(project_id=project_id, user_id=user[0], period_start=day_start, period_end=day_end, session=session)
                            difficulty = get_avg_difficulty_by_user_id_and_period(project_id=project_id, user_id=user[0], period_start=day_start, period_end=day_end, session=session)
                            cell = ws.cell(row=row_num, column=col_idx, value=f'{amount} / {difficulty}')
                            cell.font = Font(size=14)
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            total_amount += amount
                            total_difficulty += difficulty
                        cell = ws.cell(row=row_num, column=len(date_list)+2, value=f'{total_amount} / {total_difficulty}')
                        cell.font = Font(size=14)
                        cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        row_num += 1
                cell = ws.cell(row=row_num, column=1, value="Всего")
                cell.font = Font(size=14)
                cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for col_idx in range(2, len(date_list)+3):
                    total_amount = 0
                    total_difficulty = 0
                    for row_idx in range(3, row_num):
                        total_amount += int(ws.cell(row=row_idx, column=col_idx).value.split(' / ')[0])
                        total_difficulty += int(ws.cell(row=row_idx, column=col_idx).value.split(' / ')[1])
                    cell = ws.cell(row=row_num, column=col_idx, value=f'{total_amount} / {total_difficulty}')
                    cell.font = Font(size=14)
                    cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=len(date_list)+1)
            cell = ws.cell(row=1, column=2, value='Количество задач / Сумма сложностей задач')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
            # Сохраняем Excel-документ
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            headers = {"Content-Disposition": 'attachment; filename="report.xlsx"'}
            return StreamingResponse(
                excel_file,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers
            )
        elif user_id:
            users = get_teamlead_users(user_id=user_id, session=session)
            # Заполняем строки: для каждого сотрудника – подсчет выполненных задач по дням
            row_num = 3
            if users:
                user = get_id_name_by_user_id(user_id=user_id, session=session)
                cell = ws.cell(row=row_num, column=1, value=user[1])
                cell.font = Font(bold=True, size=14)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                total_amount = 0
                total_difficulty = 0
                for col_idx, single_date in enumerate(date_list, start=2):
                    day_start = datetime.combine(single_date, datetime.min.time())
                    day_end = day_start + timedelta(days=1)
                    amount = get_amount_of_done_tasks_by_user_id_and_period(user_id=user[0], period_start=day_start, period_end=day_end, session=session)
                    difficulty = get_avg_difficulty_by_user_id_and_period(user_id=user[0], period_start=day_start, period_end=day_end, session=session)
                    cell = ws.cell(row=row_num, column=col_idx, value=f'{amount} / {difficulty}')
                    cell.font = Font(size=14)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    total_amount += amount
                    total_difficulty += difficulty
                cell = ws.cell(row=row_num, column=len(date_list)+2, value=f'{total_amount} / {total_difficulty}')
                cell.font = Font(size=14)
                cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                row_num += 1
                for user in users:
                    print(user)
                    # Записываем имя сотрудника
                    cell = ws.cell(row=row_num, column=1, value=user[1])
                    cell.font = Font(size=14)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    total_amount = 0
                    total_difficulty = 0
                    for col_idx, single_date in enumerate(date_list, start=2):
                        day_start = datetime.combine(single_date, datetime.min.time())
                        day_end = day_start + timedelta(days=1)
                        amount = get_amount_of_done_tasks_by_user_id_and_period(user_id=user[0], period_start=day_start, period_end=day_end, session=session)
                        difficulty = get_avg_difficulty_by_user_id_and_period(user_id=user[0], period_start=day_start, period_end=day_end, session=session)
                        cell = ws.cell(row=row_num, column=col_idx, value=f'{amount} / {difficulty}')
                        cell.font = Font(size=14)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        total_amount += amount
                        total_difficulty += difficulty
                    cell = ws.cell(row=row_num, column=len(date_list)+2, value=f'{total_amount} / {total_difficulty}')
                    cell.font = Font(size=14)
                    cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    row_num += 1
                cell = ws.cell(row=row_num, column=1, value="Всего")
                cell.font = Font(size=14)
                cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for col_idx in range(2, len(date_list)+3):
                    total_amount = 0
                    total_difficulty = 0
                    for row_idx in range(3, row_num):
                        total_amount += int(ws.cell(row=row_idx, column=col_idx).value.split(' / ')[0])
                        total_difficulty += int(ws.cell(row=row_idx, column=col_idx).value.split(' / ')[1])
                    cell = ws.cell(row=row_num, column=col_idx, value=f'{total_amount} / {total_difficulty}')
                    cell.font = Font(size=14)
                    cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            else:
                user = get_id_name_by_user_id(user_id=user_id, session=session)
                cell = ws.cell(row=row_num, column=1, value=user[1])
                cell.font = Font(bold=True, size=14)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                total_amount = 0
                total_difficulty = 0
                for col_idx, single_date in enumerate(date_list, start=2):
                    day_start = datetime.combine(single_date, datetime.min.time())
                    day_end = day_start + timedelta(days=1)
                    amount = get_amount_of_done_tasks_by_user_id_and_period(user_id=user[0], period_start=day_start, period_end=day_end, session=session)
                    difficulty = get_avg_difficulty_by_user_id_and_period(user_id=user[0], period_start=day_start, period_end=day_end, session=session)
                    cell = ws.cell(row=row_num, column=col_idx, value=f'{amount} / {difficulty}')
                    cell.font = Font(size=14)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    total_amount += amount
                    total_difficulty += difficulty
                cell = ws.cell(row=row_num, column=len(date_list)+2, value=f'{total_amount} / {total_difficulty}')
                cell.font = Font(size=14)
                cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
                
            ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=len(date_list)+1)
            cell = ws.cell(row=1, column=2, value='Количество задач / Сумма сложностей задач')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
            # Сохраняем Excel-документ
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            headers = {"Content-Disposition": 'attachment; filename="report.xlsx"'}
            return StreamingResponse(
                excel_file,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers
            )

